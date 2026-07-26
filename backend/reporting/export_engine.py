"""
EduSense AI 360 - Export Engine
===============================

Renders a :class:`ReportData` bundle to downloadable files (Functional Requirements
Part 1B §12; Architecture Part 3 §17):

* **PDF**  - ReportLab: a formatted summary with a native engagement line chart, an
  emotion pie, the stats table, teacher insights, and student remarks.
* **Excel** - OpenPyXL: Summary, Timeline, and Students sheets.
* **CSV**  - standard library: the per-frame timeline plus a summary block.

Files are named deterministically (``<session_id>_<type>_<timestamp>.<ext>``) under
the configured export location, which is validated as writable. Failures raise
:class:`ExportError` while leaving session data intact. A simple history of generated
files is available for the Reports panel.

ReportLab/OpenPyXL are imported lazily so this module imports without them.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from config.config_manager import ConfigManager
from backend.reporting.report_engine import ReportData
from backend.contracts.models import EngagementLevel
from core.exceptions import ExportError
from core.error_handler import handle
from core.logger import get_logger
from utilities.helpers import timestamp_slug, ensure_dir

log = get_logger("application")

_EMOTION_HEX = {
    "happy": (0.98, 0.75, 0.14), "neutral": (0.58, 0.64, 0.72),
    "sad": (0.38, 0.65, 0.98), "angry": (0.97, 0.45, 0.45),
    "fear": (0.65, 0.55, 0.98), "surprise": (0.13, 0.83, 0.93),
    "disgust": (0.98, 0.57, 0.24), "confused": (0.98, 0.57, 0.24),
}


class ExportEngine:
    """Writes session reports to PDF / Excel / CSV."""

    def __init__(self, config: ConfigManager) -> None:
        self._reports_dir = Path(config.resolve_path("reports_dir"))
        self._retention = int(config.get("report.retention_count", 50))

    # -- public dispatch ----------------------------------------------------
    def export(self, data: ReportData, fmt: str) -> str:
        fmt = fmt.lower()
        if not data.has_data:
            raise ExportError("No session data to export.",
                              user_message="There is no session data to export yet.")
        dispatch = {"pdf": self.export_pdf, "excel": self.export_excel, "csv": self.export_csv}
        if fmt not in dispatch:
            raise ExportError(f"Unknown export format: {fmt}")
        return dispatch[fmt](data)

    def _target(self, session_id: str, kind: str, ext: str) -> Path:
        ensure_dir(self._reports_dir)
        if not self._reports_dir.exists():
            raise ExportError(f"Export location not writable: {self._reports_dir}")
        return self._reports_dir / f"{session_id}_{kind}_{timestamp_slug()}.{ext}"

    # -- PDF ----------------------------------------------------------------
    def export_pdf(self, data: ReportData) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.lineplots import LinePlot
            from reportlab.graphics.charts.piecharts import Pie
        except ImportError as exc:
            raise ExportError(f"ReportLab not installed: {exc}")

        s = data.session_summary
        path = self._target(s.session_id, "report", "pdf")
        try:
            doc = SimpleDocTemplate(str(path), pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
            styles = getSampleStyleSheet()
            title = ParagraphStyle("T", parent=styles["Title"], fontSize=20, spaceAfter=2)
            sub = ParagraphStyle("S", parent=styles["Normal"], fontSize=9,
                                 textColor=colors.grey, spaceAfter=12)
            story: list[Any] = [
                Paragraph(f"{data.app_name} — Session Report", title),
                Paragraph(data.app_tagline, sub),
                Paragraph(f"<b>{s.session_name}</b>", styles["Normal"]),
                Paragraph(f"Generated {data.generated_at} &nbsp;|&nbsp; Session {s.session_id}", sub),
                Spacer(1, 4 * mm),
            ]

            rows = [
                ["Metric", "Value"],
                ["Duration", f"{s.duration_seconds} s"],
                ["Average engagement", f"{s.average_engagement} / 100"],
                ["Peak / Lowest", f"{s.peak_engagement} / {s.lowest_engagement}"],
                ["Average attendance", f"{s.average_attendance} students"],
                ["Attention trend", s.attention_trend.value],
                ["Frames recorded", str(s.frames_recorded)],
            ]
            table = Table(rows, colWidths=[70 * mm, 90 * mm])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366F1")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FC")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D7DE")),
                ("FONTSIZE", (0, 0), (-1, -1), 10), ("PADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(table)
            story.append(Spacer(1, 6 * mm))

            chart = self._pdf_engagement_chart(data, Drawing, LinePlot, colors)
            if chart is not None:
                story.append(Paragraph("<b>Engagement timeline</b>", styles["Heading3"]))
                story.append(chart)
                story.append(Spacer(1, 4 * mm))

            pie = self._pdf_emotion_pie(data, Drawing, Pie, colors)
            if pie is not None:
                story.append(Paragraph("<b>Emotion distribution</b>", styles["Heading3"]))
                story.append(pie)
                story.append(Spacer(1, 4 * mm))

            story.append(Paragraph("<b>Teaching insights</b>", styles["Heading2"]))
            for o in data.teacher_observations:
                story.append(Paragraph(f"• {o}", styles["Normal"]))
            if data.teacher_suggestions:
                story.append(Spacer(1, 2 * mm))
                story.append(Paragraph("<b>Suggestions</b>", styles["Heading3"]))
                for sug in data.teacher_suggestions:
                    story.append(Paragraph(f"• {sug}", styles["Normal"]))

            if data.student_summaries:
                story.append(Spacer(1, 4 * mm))
                story.append(Paragraph("<b>Student summaries</b>", styles["Heading2"]))
                srows = [["Student", "Avg", "Trend", "Performance", "Pattern"]]
                for st in data.student_summaries:
                    srows.append([f"#{st.student_id + 1}", str(st.average_engagement),
                                  st.trend.value, st.performance, st.behaviour_pattern])
                st_table = Table(srows, colWidths=[18 * mm, 16 * mm, 24 * mm, 26 * mm, 76 * mm])
                st_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#22D3EE")),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7DE")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"), ("PADDING", (0, 0), (-1, -1), 4),
                ]))
                story.append(st_table)

            doc.build(story)
            self._prune_history()
            log.info("PDF report written: %s", path)
            return str(path)
        except ExportError:
            raise
        except Exception as exc:  # noqa: BLE001
            handle(ExportError(f"PDF export failed: {exc}"), context="pdf export", category="application")
            raise ExportError(f"PDF export failed: {exc}")

    def _pdf_engagement_chart(self, data, Drawing, LinePlot, colors):
        ts = data.timeseries
        if not ts.get("t"):
            return None
        pts = list(zip(ts["t"], ts["engagement"]))
        d = Drawing(440, 170)
        lp = LinePlot()
        lp.x, lp.y, lp.width, lp.height = 40, 20, 380, 130
        lp.data = [pts]
        lp.lines[0].strokeColor = colors.HexColor("#6366F1")
        lp.lines[0].strokeWidth = 1.5
        lp.yValueAxis.valueMin = 0
        lp.yValueAxis.valueMax = 100
        d.add(lp)
        return d

    def _pdf_emotion_pie(self, data, Drawing, Pie, colors):
        counts: dict[str, int] = {}
        for f in data.frames:
            counts[f.dominant_emotion] = counts.get(f.dominant_emotion, 0) + 1
        if not counts:
            return None
        d = Drawing(300, 160)
        pie = Pie()
        pie.x, pie.y, pie.width, pie.height = 90, 20, 120, 120
        pie.data = list(counts.values())
        pie.labels = [k.title() for k in counts]
        for i, key in enumerate(counts):
            rgb = _EMOTION_HEX.get(key, (0.58, 0.64, 0.72))
            pie.slices[i].fillColor = colors.Color(*rgb)
        d.add(pie)
        return d

    # -- Excel --------------------------------------------------------------
    def export_excel(self, data: ReportData) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as exc:
            raise ExportError(f"OpenPyXL not installed: {exc}")

        s = data.session_summary
        path = self._target(s.session_id, "report", "xlsx")
        try:
            wb = Workbook()
            head_fill = PatternFill("solid", fgColor="6366F1")
            head_font = Font(bold=True, color="FFFFFF")

            ws = wb.active
            ws.title = "Summary"
            ws["A1"] = data.app_name
            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"] = s.session_name
            summary_rows = [
                ("Metric", "Value"), ("Generated", data.generated_at),
                ("Duration (s)", s.duration_seconds), ("Average engagement", s.average_engagement),
                ("Peak engagement", s.peak_engagement), ("Lowest engagement", s.lowest_engagement),
                ("Average attendance", s.average_attendance), ("Attention trend", s.attention_trend.value),
                ("Frames recorded", s.frames_recorded),
            ]
            for i, (k, v) in enumerate(summary_rows, start=4):
                ws[f"A{i}"], ws[f"B{i}"] = k, v
                if i == 4:
                    for c in (ws[f"A{i}"], ws[f"B{i}"]):
                        c.fill, c.font = head_fill, head_font
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 24

            ts = wb.create_sheet("Timeline")
            headers = ["t (s)", "Timestamp", "Engagement", "Raw", "Faces", "Distracted", "Emotion"]
            ts.append(headers)
            for c in range(1, len(headers) + 1):
                cell = ts.cell(row=1, column=c)
                cell.fill, cell.font = head_fill, head_font
                cell.alignment = Alignment(horizontal="center")
            for f in data.frames:
                ts.append([f.t, f.timestamp, f.classroom_engagement, f.raw_engagement,
                           f.faces_present, f.distracted_count, f.dominant_emotion])
            for col in "ABCDEFG":
                ts.column_dimensions[col].width = 15

            stu = wb.create_sheet("Students")
            sh = ["Student", "Frames", "Avg", "Peak", "Lowest", "Performance",
                  "Attention %", "Trend", "Pattern"]
            stu.append(sh)
            for c in range(1, len(sh) + 1):
                cell = stu.cell(row=1, column=c)
                cell.fill, cell.font = head_fill, head_font
            for st in data.student_summaries:
                stu.append([f"#{st.student_id + 1}", st.frames, st.average_engagement,
                            st.peak_engagement, st.lowest_engagement, st.performance,
                            st.attention_ratio, st.trend.value, st.behaviour_pattern])
            for col in "ABCDEFGHI":
                stu.column_dimensions[col].width = 16

            wb.save(path)
            self._prune_history()
            log.info("Excel report written: %s", path)
            return str(path)
        except Exception as exc:  # noqa: BLE001
            handle(ExportError(f"Excel export failed: {exc}"), context="excel export", category="application")
            raise ExportError(f"Excel export failed: {exc}")

    # -- CSV ----------------------------------------------------------------
    def export_csv(self, data: ReportData) -> str:
        s = data.session_summary
        path = self._target(s.session_id, "report", "csv")
        try:
            with open(path, "w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                writer.writerow([data.app_name, "Session Report"])
                writer.writerow(["Session", s.session_name])
                writer.writerow(["Generated", data.generated_at])
                writer.writerow([])
                writer.writerow(["Summary"])
                writer.writerow(["Average engagement", s.average_engagement])
                writer.writerow(["Peak", s.peak_engagement, "Lowest", s.lowest_engagement])
                writer.writerow(["Attention trend", s.attention_trend.value])
                writer.writerow(["Average attendance", s.average_attendance])
                writer.writerow(["Frames", s.frames_recorded])
                writer.writerow([])
                writer.writerow(["t", "timestamp", "engagement", "raw", "faces", "distracted", "emotion"])
                for f in data.frames:
                    writer.writerow([f.t, f.timestamp, f.classroom_engagement, f.raw_engagement,
                                     f.faces_present, f.distracted_count, f.dominant_emotion])
            self._prune_history()
            log.info("CSV report written: %s", path)
            return str(path)
        except OSError as exc:
            handle(ExportError(f"CSV export failed: {exc}"), context="csv export", category="application")
            raise ExportError(f"CSV export failed: {exc}")

    # -- history ------------------------------------------------------------
    def history(self) -> list[dict[str, Any]]:
        """List previously generated report files (newest first)."""
        if not self._reports_dir.exists():
            return []
        items = []
        for p in self._reports_dir.iterdir():
            if p.is_file() and p.suffix.lower() in (".pdf", ".xlsx", ".csv"):
                stat = p.stat()
                items.append({
                    "name": p.name,
                    "format": p.suffix.lstrip(".").upper(),
                    "size_kb": round(stat.st_size / 1024, 1),
                    "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                    "path": str(p),
                })
        items.sort(key=lambda x: x["modified"], reverse=True)
        return items

    def _prune_history(self) -> None:
        """Keep only the most recent ``retention`` report files."""
        files = sorted(
            [p for p in self._reports_dir.iterdir() if p.is_file()
             and p.suffix.lower() in (".pdf", ".xlsx", ".csv")],
            key=lambda p: p.stat().st_mtime, reverse=True)
        for old in files[self._retention:]:
            try:
                old.unlink()
            except OSError:
                pass
